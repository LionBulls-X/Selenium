"""
============================================================
  🏢 TÜRKİYE FIRMA ARAMA BOTU
  Kaynak  : OpenStreetMap — Overpass API (ücretsiz)
  Çıktı   : Excel (.xlsx)
  Uyumlu  : Termux / Windows / Linux
============================================================
  Kurulum (Termux):
      pip install requests openpyxl

  Kullanım:
      python firma_arama.py
============================================================
"""

import requests
import json
import time
import os
import sys
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("❌ openpyxl bulunamadı. Kuruluyorum...")
    os.system(f"{sys.executable} -m pip install openpyxl")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ============================================================
# 📍 İL / İLÇE BBOX VERİLERİ
#    (enlem_min, boylam_min, enlem_max, boylam_max)
#    Siz yeni il/ilçe eklemek istediğinizde
#    Google Maps'ten koordinatları alıp buraya ekleyebilirsiniz.
# ============================================================

ILLER = {
    "İstanbul": {
        "bbox": (40.85, 28.65, 41.20, 29.25),
        "ilceler": {
            "Tüm İl":      None,  # None -> ilin tüm bbox'ı kullanılır
            "Kadıköy":     (40.97, 29.07, 41.01, 29.14),
            "Üsküdar":     (41.00, 29.05, 41.04, 29.12),
            "Beşiktaş":    (41.04, 29.00, 41.07, 29.04),
            "Şişli":       (41.05, 28.97, 41.08, 29.02),
            "Taksim/Beyoğlu":  (41.04, 28.97, 41.06, 29.00),
            "Sultanahmet/Eminönü": (41.00, 28.95, 41.02, 28.99),
            "Ataşehir":    (40.98, 29.15, 41.02, 29.22),
            "Maltepe":     (40.94, 29.10, 40.97, 29.16),
            "Pendik":      (40.91, 29.22, 40.94, 29.30),
            "Bakırköy":    (40.98, 28.85, 41.01, 28.90),
            "Zeytinburnu": (40.99, 28.88, 41.01, 28.92),
            "Esenyler":    (41.01, 28.73, 41.04, 28.77),
            "Aksaray/Merter": (40.99, 28.90, 41.02, 28.94),
            "Bahçelievler":   (40.99, 28.84, 41.02, 28.88),
            "Eyüpsultan":  (41.08, 28.93, 41.11, 28.97),
            "Fatih":       (41.02, 28.93, 41.04, 28.97),
            "Sarıyer":     (41.14, 29.08, 41.18, 29.14),
            "Beykoz":      (41.10, 29.10, 41.14, 29.18),
            "Umraniye":    (41.03, 29.08, 41.06, 29.13),
            "Çekmeköy":    (41.04, 29.12, 41.07, 29.17),
            "Sultanbeyli":  (40.99, 29.22, 41.02, 29.28),
        }
    },
    "Ankara": {
        "bbox": (39.82, 32.52, 39.97, 32.68),
        "ilceler": {
            "Tüm İl":      None,
            "Çankırı Caddesi / Kızılay": (39.91, 32.85, 39.93, 32.88),
            "Keçiören":    (39.94, 32.85, 39.97, 32.89),
            "Çankaya":     (39.90, 32.86, 39.92, 32.89),
            "Etiler/Bilkent":  (39.93, 32.84, 39.95, 32.87),
            "Mamak":       (39.88, 32.89, 39.91, 32.93),
            "Altındağ":    (39.94, 32.82, 39.96, 32.85),
            "Yenimahalle": (39.95, 32.83, 39.97, 32.86),
            "Sincan":      (39.93, 32.65, 39.95, 32.68),
            "Pursaklar":   (39.95, 32.72, 39.97, 32.75),
        }
    },
    "İzmir": {
        "bbox": (38.38, 27.00, 38.48, 27.12),
        "ilceler": {
            "Tüm İl":      None,
            "Konak":       (38.41, 27.05, 38.43, 27.08),
            "Alsancak":    (38.43, 27.05, 38.45, 27.08),
            "Karşıyaka":   (38.44, 27.05, 38.46, 27.08),
            "Bayraklı":    (38.43, 27.06, 38.45, 27.09),
            "Bornova":     (38.44, 27.08, 38.46, 27.12),
            "Güzelbahçe":  (38.40, 27.01, 38.42, 27.04),
            "Torbalı":     (38.37, 27.10, 38.39, 27.14),
        }
    },
    "Antalya": {
        "bbox": (36.87, 30.62, 36.92, 30.70),
        "ilceler": {
            "Tüm İl":      None,
            "Merkez":      (36.90, 30.64, 36.92, 30.67),
            "Lara":         (36.88, 30.68, 36.90, 30.72),
            "Konyaaltı":   (36.88, 30.62, 36.90, 30.65),
            "Muratpaşa":   (36.90, 30.65, 36.92, 30.68),
            "Döşemealtı":  (36.91, 30.63, 36.93, 30.66),
        }
    },
    "Bursa": {
        "bbox": (40.18, 29.05, 40.22, 29.12),
        "ilceler": {
            "Tüm İl":      None,
            "Merkez":      (40.19, 29.06, 40.21, 29.09),
            "Nilüfer":     (40.21, 29.05, 40.23, 29.08),
            "Osmangazi":   (40.19, 29.06, 40.21, 29.09),
            "Yıldırım":    (40.18, 29.07, 40.20, 29.10),
        }
    },
    "Adana": {
        "bbox": (36.99, 35.30, 37.02, 35.34),
        "ilceler": {
            "Tüm İl":      None,
            "Merkez":      (37.00, 35.31, 37.02, 35.33),
            "Yüreğir":     (37.01, 35.32, 37.03, 35.35),
            "Seyhan":      (37.00, 35.30, 37.02, 35.33),
        }
    },
    "Konya": {
        "bbox": (37.85, 32.48, 37.88, 32.52),
        "ilceler": {
            "Tüm İl":      None,
            "Merkez":      (37.86, 32.49, 37.88, 32.51),
            "Karatay":     (37.87, 32.48, 37.89, 32.50),
            "Selçuklu":    (37.86, 32.50, 37.88, 32.52),
        }
    },
    "Trabzon": {
        "bbox": (41.00, 39.70, 41.03, 39.74),
        "ilceler": {
            "Tüm İl":      None,
            "Merkez":      (41.01, 39.71, 41.03, 39.73),
        }
    },
    "Samsun": {
        "bbox": (41.27, 36.27, 41.30, 36.31),
        "ilceler": {
            "Tüm İl":      None,
            "Merkez":      (41.28, 36.28, 41.30, 36.30),
            "İlkadım":     (41.28, 36.28, 41.30, 36.30),
        }
    },
    "Eskişehir": {
        "bbox": (39.76, 30.52, 39.79, 30.56),
        "ilceler": {
            "Tüm İl":      None,
            "Merkez":      (39.77, 30.53, 39.79, 30.55),
            "Odunpazarı":  (39.77, 30.52, 39.79, 30.55),
            "Tepebaşı":    (39.77, 30.52, 39.79, 30.55),
        }
    },
    "Denizli": {
        "bbox": (37.78, 29.07, 37.81, 29.10),
        "ilceler": {
            "Tüm İl":      None,
            "Merkez":      (37.79, 29.08, 37.81, 29.10),
        }
    },
    "Gaziantep": {
        "bbox": (37.06, 36.37, 37.09, 36.41),
        "ilceler": {
            "Tüm İl":      None,
            "Merkez":      (37.07, 36.38, 37.09, 36.40),
            "Şahinbey":    (37.07, 36.37, 37.09, 36.39),
            "Şehitkamil":  (37.06, 36.38, 37.08, 36.41),
        }
    },
    "Şırnak": {
        "bbox": (37.05, 42.45, 37.08, 42.48),
        "ilceler": {
            "Tüm İl":      None,
            "Merkez":      (37.06, 42.46, 37.08, 42.48),
        }
    },
    "Hatay": {
        "bbox": (36.16, 36.16, 36.19, 36.20),
        "ilceler": {
            "Tüm İl":      None,
            "Antakya":     (36.16, 36.16, 36.18, 36.19),
            "İskenderun":  (36.58, 36.14, 36.61, 36.18),
        }
    },
}

# ============================================================
# 🏷️ SEKTÖR KATEGORİLERİ
#    Overpass API'nın tag sistemine göre.
#    Her sektör altında birden fazla alt-tag var —
#    hepsini aynı anda sorgulayız.
# ============================================================

SEKTORLER = {
    "Yiyecek & İçecek": {
        "amenity": ["restaurant", "cafe", "bar", "fast_food", "food_court", "pub", "ice_cream"],
        "shop":    ["food", "bakery", "butcher", "confectionery", "deli"],
    },
    "Otomotiv": {
        "amenity": ["fuel", "parking", "car_wash"],
        "shop":    ["car", "car_parts", "car_repair", "motorcycle"],
        "highway": ["service"],
    },
    "Sağlık": {
        "amenity": ["hospital", "clinic", "doctors", "dentist", "pharmacy", "veterinary"],
    },
    "Eğitim": {
        "amenity": ["school", "university", "college", "kindergarten", "library"],
        "shop":    ["education"],
    },
    "Teknoloji & Elektronik": {
        "shop":    ["electronics", "computer", "mobile_phone", "camera"],
        "amenity": ["internet_cafe"],
    },
    "İnşaat & Donanım": {
        "shop":    ["hardware", "building_materials", "garden"],
    },
    "Tekstil & Giyim": {
        "shop":    ["clothing", "shoes", "tailoring", "fabric"],
    },
    "Turizm & Konaklama": {
        "tourism": ["hotel", "motel", "hostel", "hotel", "camp_site", "apartment"],
        "amenity": ["hotel"],
    },
    "Güzellik & Kozmetik": {
        "shop":    ["beauty", "cosmetics"],
        "amenity": ["beauty"],
    },
    "Mobilya & Ev": {
        "shop":    ["furniture", "home_decoration", "kitchen", "lighting"],
    },
    "Eczane & Kozmetik": {
        "amenity": ["pharmacy"],
        "shop":    ["beauty", "cosmetics"],
    },
    "Taşımacılık & Lojistik": {
        "amenity": ["taxi", "bus_station", "ferry_terminal"],
        "shop":    ["cargo"],
    },
    "Finans & Bankacılık": {
        "amenity": ["bank", "atm", "bureau_de_change"],
    },
    "Hukuk & Danışmanlık": {
        "amenity": ["lawyers", "advice"],
    },
    "Spor & Fitness": {
        "amenity": ["sports_centre", "gym", "swimming_pool"],
        "leisure": ["sports_centre", "swimming_pool", "fitness_centre", "track"],
    },
    "Tarım & Hayvancılık": {
        "landuse": ["farm", "farmyard"],
        "shop":    ["agricultural"],
    },
    "Enerji & Solar": {
        "shop":    ["energy", "solar"],
    },
    "Kuaför & Berber": {
        "amenity": ["hairdresser"],
        "shop":    ["hairdresser"],
    },
    "Tüm Sektörler": {
        "_tum": True   # Özel bayrak — geniş sorgu yapar
    },
}

# ============================================================
# 🌐 API AYARLARI
# ============================================================

# Overpass API sunucuları (birbirinin yedek)
API_SERVERS = [
    "https://overpass-api.de/api/interpreter",
    "https://api.ubikerio.de/api/interpreter",
    "https://overpass.kumi.eu/api/interpreter",
]

REQUEST_TIMEOUT = 60  # saniye
MAX_RETRY       = 3   # her sunucu için max deneme


# ============================================================
# 🔧 YARDIMCI FONKSİYONLAR
# ============================================================

def log(msg):
    print(f"[{time.strftime('%H:%M:%S')}] {msg}")

def log_baslik(baslik):
    print("\n" + "=" * 60)
    print(f"  {baslik}")
    print("=" * 60)


def sorgu_otur(bbox, sektör_adı):
    """Overpass sorgu stringini oluştur"""

    enlem_min, boylam_min, enlem_max, boylam_max = bbox
    bbox_str = f"{enlem_min},{boylam_min},{enlem_max},{boylam_max}"

    tags = SEKTORLER[sektör_adı]

    # ── "Tüm Sektörler" modunda geniş sorgu ──
    if tags.get("_tum"):
        query = f"""
[out:json][timeout:60];
(
  node["name"](bbox);
  way["name"](bbox);
);
out center;
""".replace("bbox", bbox_str)
        return query

    # ── Normal sektör sorgusu ──
    # Her tag_type / tag_value kombinasyonu için ayrı bir blok
    blocks = []
    for tag_type, tag_values in tags.items():
        for val in tag_values:
            blocks.append(f'  node["{tag_type}"="{val}"]["name"]({bbox_str});')
            blocks.append(f'  way["{tag_type}"="{val}"]["name"]({bbox_str});')

    inner = "\n".join(blocks)

    query = f"""
[out:json][timeout:60];
(
{inner}
);
out center;
"""
    return query


def api_sorgu_yap(query):
    """Overpass API'na sorgu yap — sunucu yedekleme destekli"""
    for server in API_SERVERS:
        for attempt in range(MAX_RETRY):
            try:
                log(f"🌐 Sunucu: {server.split('//')[1].split('/')[0]} | Deneme: {attempt+1}")
                response = requests.get(
                    server,
                    params={"data": query},
                    timeout=REQUEST_TIMEOUT,
                    headers={"User-Agent": "FirmaArama/1.0 (Termux)"}
                )
                if response.status_code == 200:
                    return response.json()
                elif response.status_code == 429:
                    # Rate limit — bekle
                    log("⏳ Rate limit — 10 saniye beklenyor...")
                    time.sleep(10)
                else:
                    log(f"⚠️  HTTP {response.status_code} — sonraki sunucuya geçiyor...")
                    break  # bu sunucudan vazgeç, diğerine geç
            except requests.exceptions.Timeout:
                log(f"⏰ Timeout (deneme {attempt+1})")
                time.sleep(3)
            except requests.exceptions.ConnectionError:
                log(f"🔴 Bağlantı hatası — sonraki sunucuya geçiyor...")
                break
            except Exception as e:
                log(f"❌ Hata: {str(e)[:60]}")
                time.sleep(2)

    return None  # tüm sunucular başarısız


def parse_sonuclar(data):
    """API response → firma listesi"""
    firmalar = []
    seen = set()  # Tekrar kontrolü (ad + koordinat)

    for el in data.get("elements", []):
        tags = el.get("tags", {})
        ad   = tags.get("name", "").strip()

        if not ad:
            continue

        # Koordinat
        if el["type"] == "node":
            lat = el.get("lat")
            lon = el.get("lon")
        else:
            center = el.get("center", {})
            lat = center.get("lat")
            lon = center.get("lon")

        if not lat or not lon:
            continue

        # Tekrar kontrolü
        key = f"{ad}_{round(lat,4)}_{round(lon,4)}"
        if key in seen:
            continue
        seen.add(key)

        # Kategori tespiti
        kategori_parts = []
        for k in ["amenity", "shop", "tourism", "leisure"]:
            if k in tags:
                kategori_parts.append(tags[k])
        kategori = " / ".join(kategori_parts) if kategori_parts else "-"

        # Adres birleştirme
        sokak  = tags.get("addr:street", "-")
        no     = tags.get("addr:housenumber", "")
        adres  = f"{sokak} {no}".strip() if sokak != "-" else "-"

        firmalar.append({
            "firma_adi":      ad,
            "kategori":       kategori,
            "telefon":        tags.get("phone", "-"),
            "email":          tags.get("email", "-"),
            "adres":          adres,
            "ilce":           tags.get("addr:city", tags.get("addr:district", "-")),
            "il":             tags.get("addr:state", "-"),
            "posta_kodu":     tags.get("addr:postcode", "-"),
            "website":        tags.get("website", "-"),
            "calisma_saat":   tags.get("opening_hours", "-"),
            "lat":            lat,
            "lon":            lon,
        })

    return firmalar


# ============================================================
# 📊 EXCEL ÇIKTISI
# ============================================================

def excel_kaydet(firmalar, il_adi, ilce_adi, sektor_adi, dosya_adi):
    """Firma listesini düzgün formatlanmış Excel dosyasına kaydet"""

    wb = Workbook()
    ws = wb.active
    ws.title = "Firma Listesi"
    ws.sheet_view.showGridLines = False

    # ── Stiller ──
    title_font   = Font(name="Arial", size=15, bold=True, color="FFFFFF")
    sub_font     = Font(name="Arial", size=10, color="AAAAAA", italic=True)
    header_font  = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    data_font    = Font(name="Arial", size=10, color="333333")
    num_font     = Font(name="Arial", size=10, bold=True, color="2E5090")

    dark_fill    = PatternFill("solid", fgColor="1B2A4A")
    header_fill  = PatternFill("solid", fgColor="2E5090")
    alt_fill     = PatternFill("solid", fgColor="F0F4FA")
    white_fill   = PatternFill("solid", fgColor="FFFFFF")

    thin_border  = Border(
        left=Side(style="thin", color="D0D8E8"),
        right=Side(style="thin", color="D0D8E8"),
        top=Side(style="thin", color="D0D8E8"),
        bottom=Side(style="thin", color="D0D8E8"),
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # ── Satır 1-2: Başlık bloku ──
    ws.merge_cells("A1:L2")
    ws["A1"] = "🏢  TÜRKİYE FIRMA ARAMA SONUÇLARI"
    ws["A1"].font = title_font
    ws["A1"].fill = dark_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 18

    # ── Satır 3: Sorgu özeti ──
    ws.merge_cells("A3:L3")
    sorgu_tarihi = datetime.now().strftime("%d.%m.%Y %H:%M")
    ws["A3"] = f"İl: {il_adi}   |   İlçe: {ilce_adi}   |   Sektör: {sektor_adi}   |   Sorgu Tarihi: {sorgu_tarihi}   |   Sonuç: {len(firmalar)} firma"
    ws["A3"].font  = Font(name="Arial", size=10, bold=True, color="2E5090")
    ws["A3"].fill  = PatternFill("solid", fgColor="E8EEF7")
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 24

    # ── Satır 4: Boş ──
    ws.row_dimensions[4].height = 6

    # ── Satır 5: Sütun başlıkları ──
    headers = [
        "#", "Firma Adı", "Kategori", "Telefon", "Email",
        "Adres", "İlçe", "İl", "Posta Kodu", "Website",
        "Çalışma Saatleri", "Harita Linki"
    ]
    col_widths = [4, 30, 16, 18, 28, 28, 14, 12, 10, 32, 24, 38]

    for col_i, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=5, column=col_i, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.border    = thin_border
        cell.alignment = center
        # Sütun genişliği (A=1, B=2, ... L=12)
        col_letter = chr(64 + col_i) if col_i <= 26 else "A" + chr(64 + col_i - 26)
        ws.column_dimensions[col_letter].width = w

    ws.row_dimensions[5].height = 22

    # ── Veri satırları (6. satırdan) ──
    fields = [
        "firma_adi", "kategori", "telefon", "email",
        "adres", "ilce", "il", "posta_kodu", "website",
        "calisma_saat"
    ]

    for row_i, firma in enumerate(firmalar, 6):
        is_alt = (row_i % 2 == 0)
        fill   = alt_fill if is_alt else white_fill

        # Sıra numarası
        cell = ws.cell(row=row_i, column=1, value=row_i - 5)
        cell.font      = num_font
        cell.fill      = fill
        cell.border    = thin_border
        cell.alignment = center

        # Veri sütunları
        for col_i, field in enumerate(fields, 2):
            cell = ws.cell(row=row_i, column=col_i, value=firma[field])
            cell.font      = data_font
            cell.fill      = fill
            cell.border    = thin_border
            # Firma adı ve adres sola, diğerleri merkeze
            cell.alignment = left if col_i in [2, 6, 10, 11] else center

        # Harita linki (Google Maps deeplink)
        lat, lon = firma["lat"], firma["lon"]
        maps_link = f"https://www.google.com/maps/?q={lat},{lon}"
        cell = ws.cell(row=row_i, column=12, value=maps_link)
        cell.font      = Font(name="Arial", size=9, color="2E5090", underline="single")
        cell.fill      = fill
        cell.border    = thin_border
        cell.alignment = left
        cell.hyperlink = maps_link

        ws.row_dimensions[row_i].height = 20

    # ── Alt bilgi satırı ──
    footer_row = len(firmalar) + 7
    ws.row_dimensions[footer_row].height = 6  # boşluk

    footer_row += 1
    ws.merge_cells(f"A{footer_row}:L{footer_row}")
    ws.cell(row=footer_row, column=1,
            value=f"📌 Toplam {len(firmalar)} firma |  Kaynak: OpenStreetMap / Overpass API  |  ⚠️  Bazı bilgiler eksik olabilir.")
    ws.cell(row=footer_row, column=1).font = Font(name="Arial", size=9, italic=True, color="888888")
    ws.cell(row=footer_row, column=1).alignment = left

    # ── Kaydet ──
    wb.save(dosya_adi)
    return dosya_adi


# ============================================================
# 📋 MENÜ FONKSİYONLARI
# ============================================================

def secim_menüsü(başlık, seçenekler):
    """Numaralı menü göster, seçim al"""
    print(f"\n  📌 {başlık}")
    print("  " + "-" * 44)
    for i, s in enumerate(seçenekler, 1):
        print(f"  {i:>2}. {s}")
    print("  " + "-" * 44)

    while True:
        try:
            secim = int(input(f"\n  ➡️  Seçiminiz (1-{len(seçenekler)}): ").strip())
            if 1 <= secim <= len(seçenekler):
                return secim - 1  # 0-indexli
            print("  ❌ Geçersiz! Lütfen tekrar girin.")
        except (ValueError, EOFError):
            print("  ❌ Sayı girin lütfen.")


# ============================================================
# 🚀 ANA PROGRAM
# ============================================================

def main():
    log_baslik("🏢 TÜRKİYE FIRMA ARAMA BOTU")
    print("  Kaynak  : OpenStreetMap (Overpass API)")
    print("  Çıktı   : Excel (.xlsx)")

    # ─── 1) İL SEÇİMİ ───
    il_isimleri = list(ILLER.keys())
    il_idx      = secim_menüsü("İL SEÇİN", il_isimleri)
    il_adi      = il_isimleri[il_idx]
    log(f"✅ İl seçildi: {il_adi}")

    # ─── 2) İLÇE SEÇİMİ ───
    ilce_isimleri = list(ILLER[il_adi]["ilceler"].keys())
    ilce_idx      = secim_menüsü("İLÇE SEÇİN", ilce_isimleri)
    ilce_adi      = ilce_isimleri[ilce_idx]
    log(f"✅ İlçe seçildi: {ilce_adi}")

    # BBox belirle
    bbox = ILLER[il_adi]["ilceler"][ilce_adi]
    if bbox is None:
        bbox = ILLER[il_adi]["bbox"]  # "Tüm İl" seçildi
    log(f"📍 Sorgulanan alan (bbox): {bbox}")

    # ─── 3) SEKTÖR SEÇİMİ ───
    sektor_isimleri = list(SEKTORLER.keys())
    sektor_idx      = secim_menüsü("SEKTÖR SEÇİN", sektor_isimleri)
    sektor_adi      = sektor_isimleri[sektor_idx]
    log(f"✅ Sektör seçildi: {sektor_adi}")

    # ─── 4) Özet & Onay ───
    log_baslik("📊 SORGU ÖZETI")
    print(f"  İl      : {il_adi}")
    print(f"  İlçe    : {ilce_adi}")
    print(f"  Sektör  : {sektor_adi}")
    print(f"  BBox    : {bbox}")

    onay = input("\n  ▶️  Sorguyu başlatmak için ENTER'a basın (q = çıkış): ").strip().lower()
    if onay == "q":
        print("  👋 Çıkıldı.")
        return

    # ─── 5) API SORGUSU ───
    log_baslik("🌐 API SORGUSU YAPILIYOR...")
    query = sorgu_otur(bbox, sektor_adi)

    log("📤 Sorgu gönderildi — sonuç beklenyor...")
    start_time = time.time()

    data = api_sorgu_yap(query)

    elapsed = time.time() - start_time

    if data is None:
        log("❌ API'dan sonuç alınamadı. Bağlantınızı kontrol edin.")
        input("\n  ⏸  ENTER ile devam...")
        return

    log(f"✅ Yanıt geldi ({elapsed:.1f} sn)")

    # ─── 6) PARSE ───
    log("🔄 Sonuçlar düzenleniyor...")
    firmalar = parse_sonuclar(data)

    if not firmalar:
        log("⚠️  Hiçbir firma bulunamadı.")
        log("💡 İpucu: Farklı bir ilçe veya sektör deneyin.")
        input("\n  ⏸  ENTER ile devam...")
        return

    log(f"🏢 {len(firmalar)} firma bulundu!")

    # ─── 7) EXCEL KAYDET ───
    log("📊 Excel dosyası oluşturulyor...")

    # Dosya adı: İl_İlçe_Sektör_Tarih.xlsx
    tarih_str = datetime.now().strftime("%Y%m%d_%H%M")
    ilce_dosya = ilce_adi.replace(" ", "").replace("/", "_")
    sektor_dosya = sektor_adi.replace(" ", "").replace("&", "ve").replace("/", "_")
    dosya_adi = f"{il_adi}_{ilce_dosya}_{sektor_dosya}_{tarih_str}.xlsx"

    excel_kaydet(firmalar, il_adi, ilce_adi, sektor_adi, dosya_adi)

    log(f"✅ Excel dosyası kaydedildi:")
    log(f"   📂 {os.path.abspath(dosya_adi)}")

    # ─── 8) Kısa özet ───
    log_baslik("📋 SONUÇ ÖZETI")
    print(f"  📍 İl / İlçe   : {il_adi} / {ilce_adi}")
    print(f"  🏷️  Sektör      : {sektor_adi}")
    print(f"  🏢 Firma sayısı : {len(firmalar)}")
    print(f"  📂 Dosya        : {dosya_adi}")
    print(f"  🕐 Süre         : {elapsed:.1f} sn")

    # Telefon/email istatistik
    telefon_count = sum(1 for f in firmalar if f["telefon"] != "-")
    email_count   = sum(1 for f in firmalar if f["email"]   != "-")
    website_count = sum(1 for f in firmalar if f["website"] != "-")
    print(f"\n  📞 Telefon olan : {telefon_count} / {len(firmalar)}")
    print(f"  📧 Email olan   : {email_count} / {len(firmalar)}")
    print(f"  🌐 Website olan : {website_count} / {len(firmalar)}")

    # ─── 9) Tekrar soru ───
    tekrar = input("\n  🔄 Yeni sorgu yapmak ister misiniz? (e/h): ").strip().lower()
    if tekrar in ["e", "evet", "y", "yes"]:
        main()
    else:
        print("\n  👋 Teşekkürler! Tekrar kullanın.")


# ── Başlat ──
if __name__ == "__main__":
    main()
